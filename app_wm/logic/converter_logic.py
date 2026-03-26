#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


descricoes_remover = [
    "002 - Outros Debitos",
    "003 - Estorno de Creditos",
    "004 - Subtotal Saidas Credito do Imposto Valores Coluna Auxiliar Somas",
    "006 - Outros Creditos",
    "007 - Estorno de Debitos",
    "008 - Subtotal Entradas",
    "011 - Saldo Devedor",
    "013 - ICMS a Recolher",
    "014 - Saldo a Transportar Processamento",
    "010 - Total Entradas Apuracao do Saldo Valores Coluna Auxiliar Somas",
    "014 - Saldo a Transportar",
    "004 - Subtotal Saidas",
    "010 - Total Entradas",
    "003 - Estornos de creditos"
]

palavras_chave = [
    "002 - Outros Debitos", "002 - Outros Débitos",
    "003 - Estorno de Creditos", "003 - Estorno de Créditos",
    "004 - Subtotal Saidas", "004 - Subtotal Saídas",
    "006 - Outros Creditos", "006 - Outros Créditos",
    "007 - Estorno de Debitos", "007 - Estorno de Débitos",
    "008 - Subtotal Entradas",
    "011 - Saldo Devedor",
    "013 - ICMS a Recolher",
    "014 - Saldo a Transportar",
    "010 - Total Entradas"
]

padrao_rodape = r'Processamento:.*\d{2}/\d{2}/\d{4} \d{2}:\d{2}'
padrao_cabecalho = r'RELATÓRIO DE APOIO À APURAÇÃO DO ICMS'


def contem_valores(linha):
    return bool(re.search(r'R\$ [\d\.,]+', linha))


def converter_valor(valor):
    try:
        return float(valor.replace("R$ ", "").replace(".", "").replace(",", "."))
    except Exception:
        return valor


def extrair_dados_adicionais(texto):
    divisao = re.search(r'ESTABELECIMENTO:\s*(\d{4})', texto)
    periodo = re.search(r'(Per[ií]odo)\s*:\s*(\d{2}/\d{4})', texto)
    estabelecimento = re.search(r'ESTABELECIMENTO:\s*\d{4}\s*-\s*(.*)', texto)

    divisao = divisao.group(1) if divisao else ""
    periodo = periodo.group(2) if periodo else ""
    estabelecimento = estabelecimento.group(1).strip() if estabelecimento else ""

    estabelecimento_up = estabelecimento.upper()

    if "NORDESTE" in estabelecimento_up:
        empresa = "Gine"
    elif "NORTE" in estabelecimento_up:
        empresa = "Gino"
    else:
        empresa = "Gilda"

    return divisao, periodo, empresa


def limpar_cabecalho_grudado(descricao):
    padroes = [
        r'RELATÓRIO DE APURAÇÃO DO ICMS.*',
        r'RELATÓRIO DE APOIO À APURAÇÃO DO ICMS.*',
        r'ESTABELECIMENTO:.*',
        r'CNPJ\s*:.*',
        r'INSCR\.\s*EST\.\s*:.*',
        r'(Período|Periodo)\s*:.*',
        r'Informações de Sub-Apuração.*',
        r'Indicador de Sub-Apuração.*',
        r'Débito do Imposto.*',
        r'Crédito do Imposto.*',
        r'Apuração do Saldo.*',
        r'Valores.*',
        r'Coluna Auxiliar.*',
        r'Somas.*',
    ]
    for padrao in padroes:
        descricao = re.sub(padrao, '', descricao, flags=re.IGNORECASE).strip()
    descricao = re.sub(r'\s+', ' ', descricao).strip()
    return descricao


def processar_pdf(arquivo_pdf):
    dados = []
    linha_atual = None
    tipo_operacao = None
    ultima_linha_processada = None
    cabecalho_pagina_atual = False

    with pdfplumber.open(arquivo_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if not any(palavra in texto for palavra in palavras_chave):
                continue

            divisao, periodo, empresa = extrair_dados_adicionais(texto)
            linhas = texto.splitlines()

            for linha in linhas:
                linha = linha.strip()

                if re.match(padrao_rodape, linha):
                    continue

                if re.match(padrao_cabecalho, linha):
                    if cabecalho_pagina_atual:
                        continue
                    cabecalho_pagina_atual = True

                if "Débito do Imposto" in linha:
                    tipo_operacao = "Saida"
                elif "Crédito do Imposto" in linha:
                    tipo_operacao = "Entrada"
                elif "Apuração do Saldo" in linha:
                    tipo_operacao = "Deducao"

                if "001 - Saídas" in linha or "005 - Entradas" in linha:
                    linha = linha.replace("001 - Saídas", "NF").replace("005 - Entradas", "NF")

                valores = re.findall(r'R\$ [\d\.,]+', linha)
                descricao = re.sub(r'R\$ [\d\.,]+', '', linha).strip()
                descricao = re.sub(r"^-\s*", "", descricao)
                descricao = limpar_cabecalho_grudado(descricao)

                if contem_valores(linha):
                    if linha_atual:
                        if linha_atual != ultima_linha_processada:
                            dados.append(linha_atual)
                    linha_atual = [empresa, divisao, periodo, tipo_operacao, descricao] + valores
                else:
                    if linha_atual and descricao:
                        linha_atual[4] = limpar_cabecalho_grudado(f"{linha_atual[4]} {descricao}".strip())

            if linha_atual and linha_atual != ultima_linha_processada:
                dados.append(linha_atual)
            ultima_linha_processada = linha_atual

    return dados


def executar_conversao(pasta_pdfs, progress_callback=None):
    """
    Executa conversao completa de PDFs para Excel.
    progress_callback(pct, msg) para reportar progresso.
    Retorna lista de arquivos gerados.
    """
    arquivos_pdf = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith(".pdf")]
    if not arquivos_pdf:
        return []

    dados_compilados = []
    total = len(arquivos_pdf)

    for idx, arquivo in enumerate(arquivos_pdf):
        if progress_callback:
            progress_callback(int((idx / total) * 80), f"Processando: {arquivo}")

        arquivo_pdf = os.path.join(pasta_pdfs, arquivo)
        dados = processar_pdf(arquivo_pdf)
        dados_compilados.extend(dados)

    if not dados_compilados:
        return []

    max_valores = max(len(row) - 5 for row in dados_compilados)
    colunas = ["Empresa", "Divisao", "Periodo", "E/S", "Descricao"] + [f"Valor {i+1}" for i in range(max_valores)]
    df = pd.DataFrame(dados_compilados, columns=colunas)

    df = df[~df["Descricao"].isin(descricoes_remover)]

    for col in df.columns[5:]:
        df[col] = df[col].apply(lambda x: converter_valor(x) if isinstance(x, str) and x.startswith("R$") else x)

    arquivos_gerados = []
    grupos = list(df.groupby(["Empresa", "Periodo"], dropna=False))
    total_grupos = len(grupos)

    for idx, ((empresa, periodo), df_grupo) in enumerate(grupos):
        if progress_callback:
            progress_callback(80 + int((idx / max(total_grupos, 1)) * 20),
                            f"Exportando: {empresa} - {periodo}")

        periodo_safe = str(periodo).replace("/", ".")
        nome_arquivo = f"{empresa} - {periodo_safe}.xlsx"
        output_path = os.path.join(pasta_pdfs, nome_arquivo)

        df_grupo.to_excel(output_path, index=False, sheet_name='Consolidado')

        wb = load_workbook(output_path)
        ws = wb['Consolidado']

        moeda_style = NamedStyle(name=f"moeda_{idx}", number_format='R$ #,##0.00')

        for col in range(6, len(df_grupo.columns) + 1):
            for row in range(2, len(df_grupo) + 2):
                cell = ws.cell(row=row, column=col)
                try:
                    cell.value = float(cell.value)
                    cell.style = moeda_style
                except Exception:
                    pass

        wb.save(output_path)
        arquivos_gerados.append(output_path)

    if progress_callback:
        progress_callback(100, "Concluido!")

    return arquivos_gerados
